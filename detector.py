import cv2
from datetime import datetime

now = datetime.now()
date_string = now.strftime("%d-%m-%Y")
time_string = now.strftime("%H:%M:%S")

from openpyxl import load_workbook
from openpyxl import Workbook

# load Workbook object
wb=load_workbook("Record.xlsx")
# create new sheet
wb.create_sheet(date_string )
# select current date sheet
for s in range(len(wb.sheetnames)):
    if wb.sheetnames[s] == date_string:
        break
wb.active = s
sheet = wb.active

face_detect = cv2.CascadeClassifier('haarcascade_frontalface_default.xml')

rec=cv2.face.LBPHFaceRecognizer_create()
rec.read("recognizer\\trainingData.yml")

cascadePath = "haarcascade_frontalface_default.xml"
faceCascade = cv2.CascadeClassifier(cascadePath);

name = ''
# Ids = []
student = []

fontscale = .8
fontcolor = (255, 255, 255)

cam = cv2.VideoCapture(0)
font = cv2.FONT_HERSHEY_COMPLEX

def update_record(student):
    id =len(student)
    # set value
    sheet['A1'] = "NAME"
    sheet['B1'] = "ATTENDANCE"
    sheet['C1'] = "TIME"
    for name in student:
        sheet.cell(row=id+1, column=1).value = name
        sheet.cell(row=id+1,column=2).value = "Present"
        sheet.cell(row=id+1,column=3).value = time_string
    # save workbook
    wb.save("Record.xlsx")

while True:
    ret, img =cam.read()
    gray=cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)
    faces=faceCascade.detectMultiScale(gray, 1.2,5)
    for(x,y,w,h) in faces:
        cv2.rectangle(img,(x,y),(x+w,y+h),(225,0,0),2)
        Id, conf = rec.predict(gray[y:y+h,x:x+w])
        if(conf<50):
            if(Id==1):
                name="Samrat"
                # Ids.append(Id)
                student.append(name)
            elif(Id==2):
                name="Albina"
                # Ids.append(Id)
                student.append(name)
        else:
            name="Unknown"
        cv2.rectangle(img, (x, y + (h - 35)), (x + w, y + h), (255, 0, 0), cv2.FILLED)
        cv2.putText(img, name, (x, y + (h - 8)), font, fontscale, fontcolor, 1)
    cv2.imshow('face',img)
    if cv2.waitKey(10) & 0xFF==ord('q'):
        # Ids = (list(set(Ids)))
        student = (list(set(student)))
        # print(student)
        update_record(student)
        break
cam.release()
cv2.destroyAllWindows()
